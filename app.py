import streamlit as st
import pandas as pd
from datetime import datetime, timedelta, timezone
import io
import openpyxl
import os
import json
import gspread
from google.oauth2.service_account import Credentials

# --- 0. 사용자 및 관리자 정의, 비밀번호 세팅 ---
members = ["권회준", "김민호", "오진영", "강한수", "최지훈", "박현수", "테이"]
admins = ["장현준", "김동기", "최상철", "강택규", "김현준"]

ALL_USERS = members + admins
HOLIDAY_USERS = admins + members 

USER_PINS = {user: "5050" for user in ALL_USERS}

# 모바일/PC 넓게 쓰기 설정 및 사이드바 기본 숨김
st.set_page_config(
    page_title="T/S 근무 관리",       
    page_icon="🏢",                 
    layout="wide", 
    initial_sidebar_state="collapsed" 
)

# ⭐️ CSS 스타일 전역 주입
custom_css = """
<style>
    [data-testid="stToolbar"], [data-testid="stAppDeployButton"], 
    [data-testid="stStatusWidget"], [data-testid="stDecoration"], 
    [data-testid="collapsedControl"] { display: none !important; }
    
    .stApp, .block-container { overflow-x: hidden !important; max-width: 100vw !important; padding-top: 2rem !important; }
    
    @media (max-width: 768px) {
        h1 { white-space: nowrap !important; font-size: 5.5vw !important; letter-spacing: -0.5px !important; }
        h2 { white-space: nowrap !important; font-size: 4.5vw !important; letter-spacing: -0.5px !important; }
    }
    
    .custom-overtime-table { width: 100%; border-collapse: collapse; text-align: center; font-size: 15px; table-layout: fixed; }
    .custom-overtime-table th, .custom-overtime-table td { border: 1px solid #dcdde1; padding: 10px 2px; text-align: center !important; vertical-align: middle !important; }
    .custom-overtime-table th { background-color: #f0f2f6; color: #31333F; font-weight: bold; }
    .overtime-checked { background-color: #fff5f5; color: #ff4b4b; font-weight: bold; }
    
    @media (max-width: 768px) {
        .custom-overtime-table { font-size: 3vw !important; }
        .custom-overtime-table th, .custom-overtime-table td { padding: 4px 0px !important; height: 35px; white-space: nowrap !important; letter-spacing: -0.5px !important; }
        .overtime-checked { font-size: 2.8vw !important; letter-spacing: -1px !important; }
    }
    
    div[data-testid="stVerticalBlock"]:has(> div[data-testid="stElementContainer"] style[data-target="btn-grid"]) {
        display: flex !important; flex-direction: row !important; flex-wrap: wrap !important; gap: 8px !important;
    }
    div[data-testid="stVerticalBlock"]:has(> div[data-testid="stElementContainer"] style[data-target="btn-grid"]) > div[data-testid="stElementContainer"]:not(:has(style)) {
        width: calc(50% - 4px) !important; flex: 0 0 calc(50% - 4px) !important; min-width: 0 !important; 
    }
    div[data-testid="stVerticalBlock"]:has(> div[data-testid="stElementContainer"] style[data-target="btn-grid"]) > div[data-testid="stElementContainer"]:has(style) {
        display: none !important;
    }
    div[data-testid="stVerticalBlock"]:has(> div[data-testid="stElementContainer"] style[data-target="btn-grid"]) button {
        white-space: nowrap !important; height: auto !important; min-height: 42px !important;
    }
    
    .weekly-summary-table { width: 100%; text-align: center; font-size: 13.5px; margin-top: 10px; border-collapse: collapse; table-layout: fixed; }
    .weekly-summary-table th, .weekly-summary-table td { border: 1px solid #dcdde1; padding: 8px 4px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
    .weekly-summary-table th { background-color: #e8f0fe; color: #1a73e8; }
    .weekly-hours { font-weight: bold; color: #2c3e50; background-color: #f1f3f5; }
    .weekly-label { font-weight: bold; background-color: #f8f9fa; color: #31333F; }
    
    @media (max-width: 768px) {
        .weekly-summary-table { font-size: 2.5vw !important; }
        .weekly-summary-table th, .weekly-summary-table td { padding: 4px 2px !important; }
    }
</style>
"""
st.markdown(custom_css, unsafe_allow_html=True)

# --- 1. 세션 상태 관리 (로그인 처리) ---
if "logged_in" not in st.session_state: st.session_state.logged_in = False
if "current_user" not in st.session_state: st.session_state.current_user = None
if "login_selected_user" not in st.session_state: st.session_state.login_selected_user = ALL_USERS[0]

# 🔒 로그인 화면
if not st.session_state.logged_in:
    st.title("🏢 T/S 근무 계획 관리 시스템")
    st.caption("✨ Created by tskwon")
    
    _, col_login, _ = st.columns([1, 1.5, 1])
    with col_login:
        st.write("")
        st.markdown("### 🔐 시스템 로그인")
        
        st.markdown("##### 🧑‍💻 야근인원")
        with st.container():
            st.markdown('<style data-target="btn-grid"></style>', unsafe_allow_html=True)
            for user in members:
                btn_type = "primary" if user == st.session_state.login_selected_user else "secondary"
                if st.button(user, key=f"login_btn_{user}", use_container_width=True, type=btn_type):
                    st.session_state.login_selected_user = user
                    st.rerun()
                    
        st.write("")
        st.markdown("##### 👑 관리자")
        with st.container():
            st.markdown('<style data-target="btn-grid"></style>', unsafe_allow_html=True)
            for user in admins:
                btn_type = "primary" if user == st.session_state.login_selected_user else "secondary"
                if st.button(user, key=f"login_btn_{user}", use_container_width=True, type=btn_type):
                    st.session_state.login_selected_user = user
                    st.rerun()
        
        st.write("")
        st.markdown(f"**현재 선택됨:** `{st.session_state.login_selected_user}`")
        pin_input = st.text_input("🔑 비밀번호", type="password", placeholder="비밀번호 4자리 입력")
        
        if st.button("🚀 로그인", type="primary", use_container_width=True):
            if USER_PINS.get(st.session_state.login_selected_user) == pin_input:
                st.session_state.logged_in = True
                st.session_state.current_user = st.session_state.login_selected_user
                st.rerun()
            else:
                st.error("⚠️ 비밀번호가 일치하지 않습니다.")
    st.stop() 

# =====================================================================
# 로그인 성공 시 메인 화면
# =====================================================================
top_col1, top_col2 = st.columns([4, 1])
with top_col1:
    st.title("🏢 T/S 근무 계획 관리 시스템")
    st.caption("✨ Created by tskwon")
with top_col2:
    st.write("") 
    if st.button("🚪 로그아웃", use_container_width=True):
        st.session_state.logged_in = False
        st.session_state.current_user = None
        st.rerun()
st.write("---") 

# --- 2. 고정 데이터 및 ⭐️ 날짜(토요일) 정의 ---
night_time_slots = ["19:00", "19:30", "20:00", "20:30", "21:00", "21:30", "22:00"]
holiday_time_slots = ["12:00", "17:00"] 

KST = timezone(timedelta(hours=9))
today_date = datetime.now(KST).date()
today_str = today_date.strftime('%Y-%m-%d')

# ⭐️ 이번 주 토요일 날짜 계산 (월=0, 토=5)
this_saturday_date = today_date + timedelta(days=(5 - today_date.weekday()))
this_saturday_str = this_saturday_date.strftime('%Y-%m-%d')

# --- 3. 구글 스프레드시트 연동 ---
@st.cache_resource
def init_connection():
    key_dict = json.loads(st.secrets["gcp_service_account"])
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    creds = Credentials.from_service_account_info(key_dict, scopes=scopes)
    client = gspread.authorize(creds)
    sheet_url = "https://docs.google.com/spreadsheets/d/1v4REfMtoTB9CQzBRks45UpaVmptHxYD-mYeAOnavDvY/edit?gid=0#gid=0"
    return client.open_by_url(sheet_url).sheet1

sheet = init_connection()

# --- 4. 기타 상태 관리 ---
if "night_end_time" not in st.session_state: st.session_state.night_end_time = night_time_slots[0]
if "night_reason" not in st.session_state: st.session_state.night_reason = ""

if "holiday_end_time" not in st.session_state or st.session_state.holiday_end_time not in holiday_time_slots: 
    st.session_state.holiday_end_time = holiday_time_slots[0]
if "holiday_reason" not in st.session_state: st.session_state.holiday_reason = ""

# --- 5. 화면 레이아웃 분할 ---
col1, col2 = st.columns([1, 1.5])

# --- 왼쪽 영역: 근무 계획 등록/수정/취소 ---
with col1:
    st.header(f"📝 근무 계획 등록")
    st.markdown(f"**1. 등록 대상자:** `{st.session_state.current_user}`")
    
    tab_night, tab_holiday = st.tabs(["🌙 야간근무", "☀️ 휴일근무"])
    
    # [야간근무 탭 로직]
    with tab_night:
        st.info(f"💡 오늘(**{today_str}**) 기준으로 야근이 등록됩니다.")
        st.markdown("**2. 종료 시간을 선택하세요**")
        with st.container():
            st.markdown('<style data-target="btn-grid"></style>', unsafe_allow_html=True)
            for t_slot in night_time_slots:
                btn_type = "primary" if t_slot == st.session_state.night_end_time else "secondary"
                if st.button(t_slot, key=f"n_{t_slot}", use_container_width=True, type=btn_type):
                    st.session_state.night_end_time = t_slot
                    st.rerun()
                    
        st.write("") 
        st.markdown("**3. 근무 사유를 입력하세요 (필수)**")
        st.text_input("사유 입력", key="night_reason", label_visibility="collapsed", placeholder="예: B/T 3건 및 견뢰도 Test")
        st.write("") 

        with st.container():
            st.markdown('<style data-target="btn-grid"></style>', unsafe_allow_html=True)
            
            if st.button(f"🚀 야간 등록/수정", key="n_reg", type="primary", use_container_width=True):
                if not st.session_state.night_reason.strip():
                    st.error("⚠️ 근무 사유를 반드시 적어주세요!")
                else:
                    all_data = sheet.get_all_values()
                    row_to_update = -1
                    for i, row in enumerate(all_data):
                        row_wt = row[5] if len(row) >= 6 else "야간"
                        if i > 0 and len(row) >= 4 and row[1] == st.session_state.current_user and row[2] == today_str and row_wt == "야간":
                            row_to_update = i + 1 
                            break
                    if row_to_update != -1:
                        sheet.update_cell(row_to_update, 4, st.session_state.night_end_time) 
                        sheet.update_cell(row_to_update, 5, st.session_state.night_reason)
                        if len(all_data[row_to_update-1]) < 6: sheet.update_cell(row_to_update, 6, "야간")
                        st.success(f"🔄 야간근무 변경 완료!")
                    else:
                        new_id = len(all_data)
                        sheet.append_row([new_id, st.session_state.current_user, today_str, st.session_state.night_end_time, st.session_state.night_reason, "야간"])
                        st.success(f"🎉 야간근무 등록 완료!")
                    st.rerun()
                
            if st.button(f"🗑️ 야간 취소", key="n_del", type="secondary", use_container_width=True):
                all_data = sheet.get_all_values()
                row_to_delete = -1
                for i, row in enumerate(all_data):
                    row_wt = row[5] if len(row) >= 6 else "야간"
                    if i > 0 and len(row) >= 4 and row[1] == st.session_state.current_user and row[2] == today_str and row_wt == "야간":
                        row_to_delete = i + 1
                        break
                if row_to_delete != -1:
                    sheet.delete_rows(row_to_delete)
                    st.warning(f"🗑️ 야간근무 취소 완료!")
                else:
                    st.info(f"ℹ️ 기록 없음")
                st.rerun()

    # [휴일근무 탭 로직]
    with tab_holiday:
        # ⭐️ 토요일로 자동 지정됨을 안내
        st.info(f"💡 이번 주 토요일(**{this_saturday_str}**) 기준으로 휴일근무가 등록됩니다.")
        st.markdown("**2. 종료 시간을 선택하세요**")
        with st.container():
            st.markdown('<style data-target="btn-grid"></style>', unsafe_allow_html=True)
            for t_slot in holiday_time_slots:
                btn_type = "primary" if t_slot == st.session_state.holiday_end_time else "secondary"
                if st.button(t_slot, key=f"h_{t_slot}", use_container_width=True, type=btn_type):
                    st.session_state.holiday_end_time = t_slot
                    st.rerun()
                    
        st.write("") 
        st.markdown("**3. 근무 사유를 입력하세요 (필수)**")
        st.text_input("사유 입력", key="holiday_reason", label_visibility="collapsed", placeholder="예: 공장 라인 점검")
        st.write("") 

        with st.container():
            st.markdown('<style data-target="btn-grid"></style>', unsafe_allow_html=True)
            
            if st.button(f"☀️ 휴일 등록/수정", key="h_reg", type="primary", use_container_width=True):
                if not st.session_state.holiday_reason.strip():
                    st.error("⚠️ 근무 사유를 반드시 적어주세요!")
                else:
                    all_data = sheet.get_all_values()
                    row_to_update = -1
                    for i, row in enumerate(all_data):
                        row_wt = row[5] if len(row) >= 6 else "야간"
                        # ⭐️ today_str 대신 this_saturday_str 에 저장
                        if i > 0 and len(row) >= 4 and row[1] == st.session_state.current_user and row[2] == this_saturday_str and row_wt == "휴일":
                            row_to_update = i + 1 
                            break
                    if row_to_update != -1:
                        sheet.update_cell(row_to_update, 4, st.session_state.holiday_end_time) 
                        sheet.update_cell(row_to_update, 5, st.session_state.holiday_reason)
                        st.success(f"🔄 휴일근무 변경 완료!")
                    else:
                        new_id = len(all_data)
                        # ⭐️ today_str 대신 this_saturday_str 로 저장
                        sheet.append_row([new_id, st.session_state.current_user, this_saturday_str, st.session_state.holiday_end_time, st.session_state.holiday_reason, "휴일"])
                        st.success(f"🎉 휴일근무 등록 완료!")
                    st.rerun()
                
            if st.button(f"🗑️ 휴일 취소", key="h_del", type="secondary", use_container_width=True):
                all_data = sheet.get_all_values()
                row_to_delete = -1
                for i, row in enumerate(all_data):
                    row_wt = row[5] if len(row) >= 6 else "야간"
                    # ⭐️ 삭제할 때도 this_saturday_str 기준으로 찾음
                    if i > 0 and len(row) >= 4 and row[1] == st.session_state.current_user and row[2] == this_saturday_str and row_wt == "휴일":
                        row_to_delete = i + 1
                        break
                if row_to_delete != -1:
                    sheet.delete_rows(row_to_delete)
                    st.warning(f"🗑️ 휴일근무 취소 완료!")
                else:
                    st.info(f"ℹ️ 기록 없음")
                st.rerun()

# --- 오른쪽 영역: 탭(Tab) 기반 현황판 ---
with col2:
    view_date = st.date_input("🗓️ 조회 기준 날짜 선택 (평일을 골라도 해당 주 토요일 휴일근무 확인 가능)", today_date)
    view_str = view_date.strftime('%Y-%m-%d')
    
    # ⭐️ 조회된 날짜 기준의 토요일 날짜 계산 (휴일 탭 출력용)
    view_saturday_date = view_date + timedelta(days=(5 - view_date.weekday()))
    view_saturday_str = view_saturday_date.strftime('%Y-%m-%d')
    
    all_data = sheet.get_all_values()
    
    if st.session_state.current_user in admins:
        tabs = st.tabs(["🌙 야간근무 현황", "☀️ 휴일근무 현황 (토요일)"])
        tab1, tab2 = tabs[0], tabs[1]
        has_tab3 = False
    else:
        tab1, tab2, tab3 = st.tabs(["🌙 야간근무 현황", "☀️ 휴일근무 현황 (토요일)", "📅 나의 8주 달력 (야간+휴일)"])
        has_tab3 = True
    
    # === 탭 1: 야간근무 현황 ===
    with tab1:
        st.header(f"🌙 {view_str} 야간근무")
        grid_df = pd.DataFrame(index=night_time_slots, columns=members).fillna("")
        records_night = []
        
        for row in all_data[1:]:
            if len(row) >= 4 and row[2] == view_str:
                row_wt = row[5] if len(row) >= 6 else "야간"
                if row_wt == "야간":
                    row_name = row[1]
                    row_end_time = row[3]
                    reason = row[4] if len(row) >= 5 and row[4].strip() != "" else "업무 연장"
                    records_night.append((row_name, row_end_time, reason))
        
        records_night.sort(key=lambda x: members.index(x[0]) if x[0] in members else 999)
        
        for name, end_t, reason in records_night:
            if end_t in grid_df.index and name in grid_df.columns:
                grid_df.loc[end_t, name] = "✔️ 야근"
                    
        html_code = f'<table class="custom-overtime-table"><thead><tr><th>시간</th>'
        for col in grid_df.columns: html_code += f'<th>{col}</th>'
        html_code += '</tr></thead><tbody>'
        for index, row in grid_df.iterrows():
            html_code += f'<tr><th>{index}</th>'
            for val in row:
                html_code += f'<td class="overtime-checked">{val}</td>' if val == "✔️ 야근" else f'<td>{val}</td>'
            html_code += '</tr>'
        html_code += '</tbody></table>'
        st.markdown(html_code, unsafe_allow_html=True)
        st.write("") 
        
        template_path = "template.xlsx"
        if os.path.exists(template_path):
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            ws['F3'] = view_str
            if st.session_state.current_user in admins:
                ws['H3'] = st.session_state.current_user
            
            start_row = 8
            for idx, (name, end_t, reason) in enumerate(records_night, start=1):
                ws.cell(row=start_row, column=2, value=idx)                             
                ws.cell(row=start_row, column=3, value=name)                            
                ws.cell(row=start_row, column=5, value=f"17:30 ~ {end_t}")              
                ws.cell(row=start_row, column=6, value=reason)                      
                start_row += 1
                
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            st.download_button(
                label=f"📥 {view_str} 야간 양식 다운로드",
                data=excel_buffer.getvalue(),
                file_name=f"야근계획서_{view_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    # === 탭 2: 휴일근무 현황 ===
    with tab2:
        # ⭐️ view_str 이 아닌 view_saturday_str 을 화면에 표시
        st.header(f"☀️ {view_saturday_str} 휴일근무")
        grid_df_holiday = pd.DataFrame(index=holiday_time_slots, columns=HOLIDAY_USERS).fillna("")
        records_holiday = []
        
        for row in all_data[1:]:
            # ⭐️ view_str 이 아닌 view_saturday_str 으로 필터링하여 토요일 데이터만 뽑아옴
            if len(row) >= 4 and row[2] == view_saturday_str:
                row_wt = row[5] if len(row) >= 6 else "야간"
                if row_wt == "휴일":
                    row_name = row[1]
                    row_end_time = row[3]
                    reason = row[4] if len(row) >= 5 and row[4].strip() != "" else "휴일 특근"
                    records_holiday.append((row_name, row_end_time, reason))
        
        records_holiday.sort(key=lambda x: HOLIDAY_USERS.index(x[0]) if x[0] in HOLIDAY_USERS else 999)
        
        for name, end_t, reason in records_holiday:
            if end_t in grid_df_holiday.index and name in grid_df_holiday.columns:
                grid_df_holiday.loc[end_t, name] = "✔️ 휴일"
                    
        html_code_h = f'<table class="custom-overtime-table"><thead><tr><th>시간</th>'
        for col in grid_df_holiday.columns: html_code_h += f'<th>{col}</th>'
        html_code_h += '</tr></thead><tbody>'
        for index, row in grid_df_holiday.iterrows():
            html_code_h += f'<tr><th>{index}</th>'
            for val in row:
                html_code_h += f'<td class="overtime-checked">{val}</td>' if val == "✔️ 휴일" else f'<td>{val}</td>'
            html_code_h += '</tr>'
        html_code_h += '</tbody></table>'
        st.markdown(html_code_h, unsafe_allow_html=True)
        st.write("") 
        
        if os.path.exists(template_path):
            wb_h = openpyxl.load_workbook(template_path)
            ws_h = wb_h.active
            
            # ⭐️ 엑셀에도 토요일 날짜 출력
            ws_h['F3'] = view_saturday_str
            if st.session_state.current_user in admins:
                ws_h['H3'] = st.session_state.current_user
            
            start_row = 8
            for idx, (name, end_t, reason) in enumerate(records_holiday, start=1):
                ws_h.cell(row=start_row, column=2, value=idx)                             
                ws_h.cell(row=start_row, column=3, value=name)                            
                ws_h.cell(row=start_row, column=5, value=f"08:00 ~ {end_t}")              
                ws_h.cell(row=start_row, column=6, value=reason)                      
                start_row += 1
                
            excel_buffer_h = io.BytesIO()
            wb_h.save(excel_buffer_h)
            excel_buffer_h.seek(0)
            
            st.download_button(
                label=f"📥 {view_saturday_str} 휴일 양식 다운로드",
                data=excel_buffer_h.getvalue(),
                file_name=f"휴일근무서_{view_saturday_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    # === 탭 3: 요일별 8주 달력 (야간+휴일 합산, 일반 인원 전용) ===
    if has_tab3:
        with tab3:
            current_week_start = view_date - timedelta(days=view_date.weekday())
            weeks_info = []
            for i in range(7, -1, -1):
                w_start = current_week_start - timedelta(weeks=i)
                w_end = w_start + timedelta(days=6)
                label = f"{w_start.strftime('%m/%d')} ~ {w_end.strftime('%m/%d')}"
                weeks_info.append({"start": w_start, "end": w_end, "label": label})
                
            def calculate_work_hours(end_time_str, work_type):
                try:
                    end = datetime.strptime(end_time_str, "%H:%M")
                    if work_type == "야간":
                        start = datetime.strptime("17:30", "%H:%M")
                        hours = (end - start).total_seconds() / 3600.0
                    else: # 휴일
                        start = datetime.strptime("08:00", "%H:%M")
                        hours = (end - start).total_seconds() / 3600.0
                        if end > datetime.strptime("12:00", "%H:%M"):
                            hours -= 1.0
                            
                    return max(hours, 0.0) 
                except ValueError:
                    return 0.0

            st.subheader("📅 나의 8주 상세 달력")
            st.info(f"이 데이터는 오직 **{st.session_state.current_user}** 님에게만 표시됩니다.")
            target_user = st.session_state.current_user

            calendar_data = { w["label"]: [0.0] * 7 for w in weeks_info }

            for row in all_data[1:]: 
                if len(row) >= 4:
                    try:
                        row_name = row[1]
                        if row_name == target_user:
                            row_date = datetime.strptime(row[2], "%Y-%m-%d").date()
                            row_end_time = row[3]
                            row_wt = row[5] if len(row) >= 6 else "야간"
                            
                            if weeks_info[0]["start"] <= row_date <= weeks_info[-1]["end"]:
                                for w in weeks_info:
                                    if w["start"] <= row_date <= w["end"]:
                                        if row_wt == "휴일":
                                            day_idx = 5
                                        else:
                                            day_idx = row_date.weekday()
                                            
                                        calendar_data[w["label"]][day_idx] += calculate_work_hours(row_end_time, row_wt)
                                        break
                    except ValueError:
                        continue
            
            weekly_html = '''
            <table class="weekly-summary-table">
                <colgroup>
                    <col style="width: 26%;">
                    <col style="width: 10%;">
                    <col style="width: 10%;">
                    <col style="width: 10%;">
                    <col style="width: 10%;">
                    <col style="width: 10%;">
                    <col style="width: 10%;">
                    <col style="width: 14%;">
                </colgroup>
                <thead>
                    <tr>
                        <th>주차 (기간)</th>
                        <th>월</th><th>화</th><th>수</th><th>목</th><th>금</th><th>토</th>
                        <th>합계</th>
                    </tr>
                </thead>
                <tbody>
            '''
            
            for w in weeks_info:
                label = w["label"]
                days = calendar_data[label][:6] 
                week_total = sum(days)
                
                weekly_html += f'<tr><td class="weekly-label">{label}</td>'
                
                for d in days:
                    display_d = f"{d:.1f}h" if d > 0 else "-"
                    weekly_html += f'<td>{display_d}</td>'
                    
                display_total = f"{week_total:.1f}h" if week_total > 0 else "-"
                weekly_html += f'<td class="weekly-hours">{display_total}</td></tr>'
                
            weekly_html += '</tbody></table>'
            st.markdown(weekly_html, unsafe_allow_html=True)